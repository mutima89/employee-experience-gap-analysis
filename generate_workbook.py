#!/usr/bin/env python3
"""
Employee Experience Gap Analysis & Retention Risk Assessment
Workbook Generator

Generates a professional Excel workbook with:
  - Survey design (30 items, 6 dimensions)
  - Raw data entry with sample data (50 respondents)
  - Scoring engine with native Excel formulas
  - Executive dashboard (KPIs, radar chart, heat map, risk matrix)
  - Cronbach's Alpha reliability analysis
  - Driver analysis (correlation)
  - Turnover prediction model

All calculations use native Excel formulas (no VBA).
"""

import random
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import RadarChart, Reference, ScatterChart, Series, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURATION
# ============================================================
import os, tempfile, shutil
TEMP_FILE = os.path.join(tempfile.gettempdir(),
                         "Employee-Experience-Gap-Analysis.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "employee-experience-gap-analysis.xlsx")
PROJECT_TITLE = "Employee Experience Gap Analysis & Retention Risk Assessment"
NUM_RESPONDENTS = 50
MAX_DATA_ROWS = 500

# Seed for reproducibility
RANDOM_SEED = 42
random.seed(RANDOM_SEED)

# Color Palette
NAVY       = "1F4E79"
BLUE       = "2E75B6"
LIGHT_BLUE = "D6E4F0"
TEAL       = "00B0F0"
GREEN      = "00B050"
LIGHT_GREEN= "E2EFDA"
AMBER      = "FFC000"
LIGHT_AMBER= "FFF2CC"
ORANGE     = "ED7D31"
LIGHT_ORANGE="FCE4D6"
RED        = "C00000"
LIGHT_RED  = "FCE4D6"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MID_GRAY   = "D9D9D9"
DARK_GRAY  = "404040"

# Styles
navy_fill   = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
blue_fill   = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
lt_blue_fill= PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
green_fill  = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
lt_green_fill=PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
amber_fill  = PatternFill(start_color=AMBER, end_color=AMBER, fill_type='solid')
lt_amber_fill=PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type='solid')
orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
lt_orange_fill=PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type='solid')
red_fill    = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
lt_red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')
white_fill  = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
lt_gray_fill= PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
md_gray_fill= PatternFill(start_color=MID_GRAY, end_color=MID_GRAY, fill_type='solid')

title_font    = Font(name='Calibri', bold=True, color=WHITE,   size=16)
h1_font       = Font(name='Calibri', bold=True, color=NAVY,    size=14)
h2_font       = Font(name='Calibri', bold=True, color=BLUE,    size=12)
h3_font       = Font(name='Calibri', bold=True, color=DARK_GRAY, size=11)
header_font   = Font(name='Calibri', bold=True, color=WHITE,   size=10)
data_font     = Font(name='Calibri', color=DARK_GRAY, size=10)
bold_data_font= Font(name='Calibri', bold=True, color=DARK_GRAY, size=10)
kpi_value_font= Font(name='Calibri', bold=True, color=WHITE,   size=22)
small_font    = Font(name='Calibri', color=DARK_GRAY, size=9)

thin_border   = Border(
    left=Side(style='thin', color=MID_GRAY),
    right=Side(style='thin', color=MID_GRAY),
    top=Side(style='thin', color=MID_GRAY),
    bottom=Side(style='thin', color=MID_GRAY)
)
header_border = Border(
    left=Side(style='thin', color=NAVY),
    right=Side(style='thin', color=NAVY),
    top=Side(style='thin', color=NAVY),
    bottom=Side(style='medium', color=NAVY)
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

# Dimension Definitions
DIMENSIONS = [
    {
        "name": "Job Fit",
        "abbr": "JF",
        "weight": 0.20,
        "items": [
            ("Role Clarity",             "My role and responsibilities are clearly defined"),
            ("Skill Utilization",        "My skills and strengths are fully utilized in my work"),
            ("Career Growth Alignment",  "There is a clear career path aligned with my aspirations"),
            ("Workload Appropriateness", "My workload is manageable and realistic"),
            ("Meaningfulness of Work",   "My work feels meaningful and purposeful"),
        ]
    },
    {
        "name": "Manager Effectiveness",
        "abbr": "ME",
        "weight": 0.20,
        "items": [
            ("Coaching Quality",    "My manager provides effective coaching and guidance"),
            ("Feedback Quality",    "I receive regular, constructive feedback from my manager"),
            ("Fairness",           "My manager treats everyone fairly and without bias"),
            ("Communication",      "My manager communicates clearly and transparently"),
            ("Trust",              "I trust my manager's decisions and judgment"),
        ]
    },
    {
        "name": "Team Collaboration",
        "abbr": "TC",
        "weight": 0.15,
        "items": [
            ("Cooperation",        "Team members cooperate effectively to achieve goals"),
            ("Psychological Safety","I feel safe expressing opinions without fear of negative consequences"),
            ("Knowledge Sharing",  "Knowledge and information are freely shared across the team"),
            ("Conflict Resolution","Disagreements are resolved constructively and respectfully"),
            ("Supportiveness",     "Team members actively support one another"),
        ]
    },
    {
        "name": "Workplace Culture",
        "abbr": "WC",
        "weight": 0.15,
        "items": [
            ("Inclusion",          "I feel included and belong in this organization"),
            ("Respect",            "People treat each other with respect at all levels"),
            ("Transparency",       "Organizational decisions are communicated transparently"),
            ("Recognition",        "Contributions are recognized and appreciated"),
            ("Values Alignment",   "The company's values align with my personal values"),
        ]
    },
    {
        "name": "Innovation & Empowerment",
        "abbr": "IE",
        "weight": 0.15,
        "items": [
            ("Autonomy",              "I have the autonomy to decide how to do my work"),
            ("Decision Authority",    "I have authority to make decisions within my role"),
            ("Idea Acceptance",       "New ideas are welcomed and seriously considered"),
            ("Learning Opportunities","I have access to learning and development opportunities"),
            ("Experimentation Support","The organization supports experimentation and calculated risk-taking"),
        ]
    },
    {
        "name": "Rewards & Growth",
        "abbr": "RG",
        "weight": 0.15,
        "items": [
            ("Compensation Fairness", "My compensation is fair relative to my role and contributions"),
            ("Promotion Opportunities","There are fair opportunities for promotion and advancement"),
            ("Recognition",           "Achievements are recognized and rewarded appropriately"),
            ("Training",              "I receive adequate training to perform and grow in my role"),
            ("Career Progression",    "I can see a clear path for career progression here"),
        ]
    }
]

# Sample Data Generation Parameters
DIM_PARAMS = {
    "Job Fit":                 {"p": 7.8, "e": 6.2, "ps": 1.2, "es": 1.5},
    "Manager Effectiveness":   {"p": 8.2, "e": 6.5, "ps": 1.1, "es": 1.6},
    "Team Collaboration":      {"p": 7.5, "e": 6.5, "ps": 1.2, "es": 1.4},
    "Workplace Culture":       {"p": 8.0, "e": 5.5, "ps": 1.0, "es": 1.7},
    "Innovation & Empowerment":{"p": 7.5, "e": 5.5, "ps": 1.3, "es": 1.6},
    "Rewards & Growth":        {"p": 8.0, "e": 5.0, "ps": 1.1, "es": 1.8},
}

DEPARTMENTS = ["Engineering", "Marketing", "Sales", "Human Resources",
               "Finance", "Operations"]
LOCATIONS = ["New York", "London", "Tokyo", "Sydney", "Berlin"]
BUSINESS_UNITS = ["Core Product", "Growth", "Enterprise", "Corporate"]
MANAGERS = ["M. Chen", "A. Patel", "S. Johnson", "L. Williams",
            "R. Kim", "T. Garcia", "J. Brown", "K. Taylor"]


# ============================================================
# HELPER FUNCTIONS
# ============================================================
def clamp(val, lo=1, hi=10):
    return max(lo, min(hi, round(val)))

def generate_score(base, std):
    return clamp(random.gauss(base, std))

def raw_data_col_letter(global_item_idx, is_promise=True):
    """
    Raw Data columns: A=RespID, B=Dept, C=Location, D=BU, E=Manager,
                      F=P_00, G=E_00, H=P_01, I=E_01, ...
    """
    col_idx = 6 + global_item_idx * 2 + (0 if is_promise else 1)
    return get_column_letter(col_idx)

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = header_border

def set_col_widths(ws, widths_dict):
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


# ============================================================
# WORKBOOK GENERATION
# ============================================================
wb = Workbook()

# ===== SHEET 1: SURVEY =====================================
ws_survey = wb.active
ws_survey.title = "Survey"
ws_survey.sheet_properties.tabColor = NAVY

ws_survey.merge_cells('A1:F1')
title_cell = ws_survey['A1']
title_cell.value = "Employee Experience Survey -- Item Bank"
title_cell.font = h1_font
title_cell.alignment = Alignment(horizontal='left', vertical='center')
ws_survey.row_dimensions[1].height = 30

ws_survey.merge_cells('A2:F2')
ws_survey['A2'].value = ("Each item rated twice: "
    "Promise Score (1-10) = 'To what extent was this promised/expected?' | "
    "Experience Score (1-10) = 'To what extent do you experience this today?'")
ws_survey['A2'].font = Font(name='Calibri', italic=True, color=DARK_GRAY, size=9)
ws_survey.row_dimensions[2].height = 22

headers = ["Dimension", "Item #", "Measure", "Promise Question",
           "Experience Question", "Weight"]
for c, h in enumerate(headers, 1):
    ws_survey.cell(row=4, column=c, value=h)
apply_header_style(ws_survey, 4, 6)
ws_survey.row_dimensions[4].height = 22

row = 5
item_global = 1

for dim in DIMENSIONS:
    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type='solid')
    for item_idx, (measure, question) in enumerate(dim["items"]):
        ws_survey.cell(row=row, column=1, value=dim["name"])
        ws_survey.cell(row=row, column=2, value=item_global)
        ws_survey.cell(row=row, column=3, value=measure)
        ws_survey.cell(row=row, column=4,
                       value=f"To what extent was '{measure}' promised, communicated, "
                             f"or expected by the organization?")
        ws_survey.cell(row=row, column=5,
                       value=f"To what extent do you actually experience "
                             f"'{measure}' in your work today?")
        ws_survey.cell(row=row, column=6, value=dim["weight"])

        for c in range(1, 7):
            cell = ws_survey.cell(row=row, column=c)
            cell.font = data_font
            cell.alignment = left_align if c >= 3 else center_align
            cell.border = thin_border
            if dim["name"] in ("Team Collaboration", "Innovation & Empowerment"):
                cell.fill = alt_fill
        ws_survey.cell(row=row, column=6).number_format = '0%'
        row += 1
        item_global += 1
    row += 1

set_col_widths(ws_survey, {'A': 22, 'B': 8, 'C': 24, 'D': 55, 'E': 55, 'F': 10})

row += 1
ws_survey.merge_cells(f'A{row}:F{row}')
ws_survey.cell(row=row, column=1,
    value="Scale: 1 = 'Not at all' | 10 = 'Completely/To a great extent'").font = small_font
row += 1
ws_survey.merge_cells(f'A{row}:F{row}')
ws_survey.cell(row=row, column=1,
    value="Dimensions: 6 dimensions x 5 items = 30 items total | "
          "Weights: Job Fit(20%), Manager(20%), Team(15%), Culture(15%), "
          "Innovation(15%), Rewards(15%)").font = small_font

ws_survey.freeze_panes = 'A5'

print("[OK] Sheet 1: Survey -- complete")


# ===== SHEET 2: RAW DATA ===================================
ws_data = wb.create_sheet("Raw Data")
ws_data.sheet_properties.tabColor = BLUE

ws_data.merge_cells('A1:BM1')
ws_data['A1'].value = "Raw Survey Data -- Respondent Scores"
ws_data['A1'].font = h1_font
ws_data.row_dimensions[1].height = 28

def generate_respondent_data():
    dept = random.choice(DEPARTMENTS)
    loc  = random.choice(LOCATIONS)
    bu   = random.choice(BUSINESS_UNITS)
    mgr  = random.choice(MANAGERS)

    scores = {}
    for dim in DIMENSIONS:
        params = DIM_PARAMS[dim["name"]]
        for idx in range(len(dim["items"])):
            p = generate_score(params["p"], params["ps"])
            e = generate_score(params["e"], params["es"])
            if random.random() < 0.05:
                p = random.randint(9, 10)
                e = random.randint(1, 3)
            scores[(dim["name"], idx)] = (p, e)
    return dept, loc, bu, mgr, scores

respondents = []
for _ in range(NUM_RESPONDENTS):
    respondents.append(generate_respondent_data())

data_headers = ["RespID", "Department", "Location", "Business Unit", "Manager"]
for dim_idx, dim in enumerate(DIMENSIONS):
    for item_idx in range(len(dim["items"])):
        global_i = dim_idx * 5 + item_idx
        data_headers.append(f"P{global_i+1:02d}")
        data_headers.append(f"E{global_i+1:02d}")

for c, h in enumerate(data_headers, 1):
    ws_data.cell(row=3, column=c, value=h)
apply_header_style(ws_data, 3, len(data_headers))
ws_data.row_dimensions[3].height = 20

for r_idx, (dept, loc, bu, mgr, scores) in enumerate(respondents):
    row = 4 + r_idx
    ws_data.cell(row=row, column=1, value=f"R{r_idx+1:03d}")
    ws_data.cell(row=row, column=2, value=dept)
    ws_data.cell(row=row, column=3, value=loc)
    ws_data.cell(row=row, column=4, value=bu)
    ws_data.cell(row=row, column=5, value=mgr)

    col = 6
    for dim_idx, dim in enumerate(DIMENSIONS):
        for item_idx in range(len(dim["items"])):
            p, e = scores.get((dim["name"], item_idx), (5, 5))
            ws_data.cell(row=row, column=col, value=p)
            ws_data.cell(row=row, column=col+1, value=e)
            col += 2

for r in range(4, 4 + NUM_RESPONDENTS):
    for c in range(1, len(data_headers) + 1):
        cell = ws_data.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = center_align if c > 5 else left_align
        cell.border = thin_border
        if r % 2 == 0:
            cell.fill = lt_gray_fill

col_widths = {'A': 10, 'B': 18, 'C': 14, 'D': 18, 'E': 16}
for c in range(6, len(data_headers) + 1):
    col_widths[get_column_letter(c)] = 6
set_col_widths(ws_data, col_widths)

ws_data.freeze_panes = 'G4'

dv_score = DataValidation(type="whole", operator="between", formula1=1, formula2=10)
dv_score.error = "Score must be between 1 and 10"
dv_score.errorTitle = "Invalid Score"
for c in range(6, len(data_headers) + 1):
    col_letter = get_column_letter(c)
    dv_score.add(f"{col_letter}4:{col_letter}{4+NUM_RESPONDENTS-1}")
ws_data.add_data_validation(dv_score)

print("[OK] Sheet 2: Raw Data -- complete")


# ===== SHEET 3: SCORING ENGINE =============================
ws_score = wb.create_sheet("Scoring Engine")
ws_score.sheet_properties.tabColor = GREEN

ws_score.merge_cells('A1:H1')
ws_score['A1'].value = "Scoring Engine -- Dimension & Overall Calculations"
ws_score['A1'].font = h1_font
ws_score.row_dimensions[1].height = 28

ws_score.merge_cells('A2:H2')
ws_score['A2'].value = "All calculations use native Excel formulas. Extend row references as you add respondents."
ws_score['A2'].font = Font(name='Calibri', italic=True, color=DARK_GRAY, size=9)

summary_headers = ["Dimension", "Avg Promise", "Avg Experience",
                   "Gap", "Health Score (0-100)", "Classification", "Weight"]
for c, h in enumerate(summary_headers, 1):
    ws_score.cell(row=4, column=c, value=h)
apply_header_style(ws_score, 4, len(summary_headers))

dr_start = 4
dr_end   = 4 + NUM_RESPONDENTS - 1

for d_idx, dim in enumerate(DIMENSIONS):
    row = 5 + d_idx
    ws_score.cell(row=row, column=1, value=dim["name"])

    promise_refs = []
    exp_refs     = []
    for item_idx in range(len(dim["items"])):
        global_i = d_idx * 5 + item_idx
        p_col = raw_data_col_letter(global_i, True)
        e_col = raw_data_col_letter(global_i, False)
        promise_refs.append(f"'Raw Data'!{p_col}{dr_start}:{p_col}{dr_end}")
        exp_refs.append(f"'Raw Data'!{e_col}{dr_start}:{e_col}{dr_end}")

    promise_formula = f"=AVERAGE({','.join(promise_refs)})"
    ws_score.cell(row=row, column=2, value=promise_formula)
    ws_score.cell(row=row, column=2).number_format = '0.00'

    exp_formula = f"=AVERAGE({','.join(exp_refs)})"
    ws_score.cell(row=row, column=3, value=exp_formula)
    ws_score.cell(row=row, column=3).number_format = '0.00'

    ws_score.cell(row=row, column=4,
                  value=f"=B{row}-C{row}")
    ws_score.cell(row=row, column=4).number_format = '0.00'

    ws_score.cell(row=row, column=5,
                  value=f"=MAX(0,MIN(100,100-D{row}*20))")
    ws_score.cell(row=row, column=5).number_format = '0.0'

    ws_score.cell(row=row, column=6,
                  value=f'=IF(D{row}<=1,"Healthy",IF(D{row}<=2,"Watch",'
                         f'IF(D{row}<=3,"Concern","High Risk")))')

    ws_score.cell(row=row, column=7, value=dim["weight"])
    ws_score.cell(row=row, column=7).number_format = '0%'

idx_row = 11
cell_a = ws_score.cell(row=idx_row, column=1,
              value="Overall Experience Index")
cell_a.font = bold_data_font
cell_a.fill = lt_blue_fill
cell_a.border = thin_border
cell_b = ws_score.cell(row=idx_row, column=2,
              value=f"=SUMPRODUCT(E5:E10,G5:G10)/SUM(G5:G10)")
cell_b.number_format = '0.0'
cell_b.font = Font(name='Calibri', bold=True, color=NAVY, size=14)
cell_b.fill = lt_blue_fill
cell_b.alignment = center_align
cell_b.border = thin_border
for c in range(3, 8):
    ws_score.cell(row=idx_row, column=c).fill = lt_blue_fill
    ws_score.cell(row=idx_row, column=c).border = thin_border

risk_row = 12
cell_a = ws_score.cell(row=risk_row, column=1,
              value="Retention Risk Score")
cell_a.font = bold_data_font
cell_a.fill = lt_red_fill
cell_a.border = thin_border
cell_b = ws_score.cell(row=risk_row, column=2,
              value=f"=(D6*0.25)+(D8*0.20)+(D10*0.20)+(D5*0.15)+(D7*0.10)+(D9*0.10)")
cell_b.number_format = '0.0'
cell_b.font = Font(name='Calibri', bold=True, color=RED, size=14)
cell_b.fill = lt_red_fill
cell_b.alignment = center_align
cell_b.border = thin_border
for c in range(3, 8):
    ws_score.cell(row=risk_row, column=c).fill = lt_red_fill
    ws_score.cell(row=risk_row, column=c).border = thin_border

risk_class_row = 13
cell_a = ws_score.cell(row=risk_class_row, column=1,
              value="Retention Risk Classification")
cell_a.font = bold_data_font
cell_a.fill = lt_red_fill
cell_a.border = thin_border
cell_b = ws_score.cell(row=risk_class_row, column=2,
              value=f'=IF(B12<=20,"Very Low Risk",'
                    f'IF(B12<=40,"Low Risk",'
                    f'IF(B12<=60,"Moderate Risk",'
                    f'IF(B12<=80,"High Risk","Critical Risk"))))')
cell_b.font = Font(name='Calibri', bold=True, color=RED, size=11)
cell_b.fill = lt_red_fill
cell_b.alignment = center_align
cell_b.border = thin_border
for c in range(3, 8):
    ws_score.cell(row=risk_class_row, column=c).fill = lt_red_fill
    ws_score.cell(row=risk_class_row, column=c).border = thin_border

for r in range(5, 11):
    for c in range(1, 8):
        cell = ws_score.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        if r % 2 == 0:
            cell.fill = lt_gray_fill

ws_score.conditional_formatting.add(
    'F5:F10',
    CellIsRule(operator='equal', formula=['"Healthy"'],
               fill=lt_green_fill, font=Font(color=GREEN, bold=True))
)
ws_score.conditional_formatting.add(
    'F5:F10',
    CellIsRule(operator='equal', formula=['"Watch"'],
               fill=lt_amber_fill, font=Font(color=AMBER, bold=True))
)
ws_score.conditional_formatting.add(
    'F5:F10',
    CellIsRule(operator='equal', formula=['"Concern"'],
               fill=lt_orange_fill, font=Font(color=ORANGE, bold=True))
)
ws_score.conditional_formatting.add(
    'F5:F10',
    CellIsRule(operator='equal', formula=['"High Risk"'],
               fill=lt_red_fill, font=Font(color=RED, bold=True))
)

# Per-Respondent Dimension Scores
section_b_start = 15
ws_score.merge_cells(f'A{section_b_start}:R{section_b_start}')
ws_score.cell(row=section_b_start, column=1,
              value="Per-Respondent Dimension Scores & Overall Index").font = h2_font
ws_score.row_dimensions[section_b_start].height = 22

b_header_row = section_b_start + 1
b_headers = ["RespID", "Department", "Location", "Business Unit", "Manager"]
for dim in DIMENSIONS:
    b_headers.append(f"{dim['abbr']}_P_Avg")
    b_headers.append(f"{dim['abbr']}_E_Avg")
    b_headers.append(f"{dim['abbr']}_Gap")
b_headers += ["Overall_P", "Overall_E", "Overall_Gap", "Retention_Risk", "Turnover_Prob%"]

for c, h in enumerate(b_headers, 1):
    ws_score.cell(row=b_header_row, column=c, value=h)
apply_header_style(ws_score, b_header_row, len(b_headers))

max_formula_row = b_header_row + MAX_DATA_ROWS
rd_start = 4

for resp_offset in range(MAX_DATA_ROWS):
    row = b_header_row + 1 + resp_offset
    rd_row = rd_start + resp_offset

    ws_score.cell(row=row, column=1,
                  value=f'=IF(\'Raw Data\'!A{rd_row}<>"",\'Raw Data\'!A{rd_row},"")')
    ws_score.cell(row=row, column=2,
                  value=f'=IF(\'Raw Data\'!B{rd_row}<>"",\'Raw Data\'!B{rd_row},"")')
    ws_score.cell(row=row, column=3,
                  value=f'=IF(\'Raw Data\'!C{rd_row}<>"",\'Raw Data\'!C{rd_row},"")')
    ws_score.cell(row=row, column=4,
                  value=f'=IF(\'Raw Data\'!D{rd_row}<>"",\'Raw Data\'!D{rd_row},"")')
    ws_score.cell(row=row, column=5,
                  value=f'=IF(\'Raw Data\'!E{rd_row}<>"",\'Raw Data\'!E{rd_row},"")')

    col = 6
    for d_idx, dim in enumerate(DIMENSIONS):
        p_refs = []
        e_refs = []
        for item_idx in range(len(dim["items"])):
            global_i = d_idx * 5 + item_idx
            p_col = raw_data_col_letter(global_i, True)
            e_col = raw_data_col_letter(global_i, False)
            p_refs.append(f"'Raw Data'!{p_col}{rd_row}")
            e_refs.append(f"'Raw Data'!{e_col}{rd_row}")

        ws_score.cell(row=row, column=col,
                      value=f'=IF(\'Raw Data\'!A{rd_row}<>"",'
                            f'AVERAGE({",".join(p_refs)}),"")')
        ws_score.cell(row=row, column=col).number_format = '0.00'
        col += 1

        ws_score.cell(row=row, column=col,
                      value=f'=IF(\'Raw Data\'!A{rd_row}<>"",'
                            f'AVERAGE({",".join(e_refs)}),"")')
        ws_score.cell(row=row, column=col).number_format = '0.00'
        col += 1

        p_col_letter = get_column_letter(col - 2)
        e_col_letter = get_column_letter(col - 1)
        ws_score.cell(row=row, column=col,
                      value=f'=IF({p_col_letter}{row}<>"",'
                            f'{p_col_letter}{row}-{e_col_letter}{row},"")')
        ws_score.cell(row=row, column=col).number_format = '0.00'
        col += 1

    overall_p_cols = []
    for d_idx in range(6):
        overall_p_cols.append(f"{get_column_letter(6 + d_idx*3)}{row}")
    ws_score.cell(row=row, column=col,
                  value=f'=IF(A{row}<>"",AVERAGE({",".join(overall_p_cols)}),"")')
    ws_score.cell(row=row, column=col).number_format = '0.00'
    overall_p_col_letter = get_column_letter(col)
    col += 1

    overall_e_cols = []
    for d_idx in range(6):
        overall_e_cols.append(f"{get_column_letter(6 + d_idx*3 + 1)}{row}")
    ws_score.cell(row=row, column=col,
                  value=f'=IF(A{row}<>"",AVERAGE({",".join(overall_e_cols)}),"")')
    ws_score.cell(row=row, column=col).number_format = '0.00'
    overall_e_col_letter = get_column_letter(col)
    col += 1

    ws_score.cell(row=row, column=col,
                  value=f'=IF(A{row}<>"",'
                        f'{overall_p_col_letter}{row}-{overall_e_col_letter}{row},"")')
    ws_score.cell(row=row, column=col).number_format = '0.00'
    col += 1

    jf_gap  = f"{get_column_letter(8)}{row}"
    me_gap  = f"{get_column_letter(11)}{row}"
    tc_gap  = f"{get_column_letter(14)}{row}"
    wc_gap  = f"{get_column_letter(17)}{row}"
    ie_gap  = f"{get_column_letter(20)}{row}"
    rg_gap  = f"{get_column_letter(23)}{row}"

    ws_score.cell(row=row, column=col,
                  value=f'=IF(A{row}<>"",'
                        f'({me_gap}*0.25)+({wc_gap}*0.20)+({rg_gap}*0.20)'
                        f'+({jf_gap}*0.15)+({tc_gap}*0.10)+({ie_gap}*0.10),"")')
    ws_score.cell(row=row, column=col).number_format = '0.00'
    risk_col_letter = get_column_letter(col)
    col += 1

    ws_score.cell(row=row, column=col,
                  value=f'=IF(A{row}<>"",'
                        f'({risk_col_letter}{row}*0.5)+(0.3*{me_gap})+(0.2*{wc_gap}),"")')
    ws_score.cell(row=row, column=col).number_format = '0.0%'

data_end = b_header_row + MAX_DATA_ROWS
for r in range(b_header_row + 1, data_end + 1):
    for c in range(1, len(b_headers) + 1):
        cell = ws_score.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        if r % 2 == 0:
            cell.fill = lt_gray_fill

score_widths = {'A': 10, 'B': 18, 'C': 14, 'D': 18, 'E': 16}
for i in range(6, len(b_headers) + 1):
    score_widths[get_column_letter(i)] = 11
set_col_widths(ws_score, score_widths)

ws_score.freeze_panes = f'F{b_header_row + 1}'

print("[OK] Sheet 3: Scoring Engine -- complete")


# ===== SHEET 4: DASHBOARD ==================================
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = BLUE

ws_dash.merge_cells('A1:R1')
ws_dash['A1'].value = PROJECT_TITLE
ws_dash['A1'].font = title_font
ws_dash['A1'].fill = navy_fill
ws_dash['A1'].alignment = Alignment(horizontal='left', vertical='center')
for c in range(1, 19):
    ws_dash.cell(row=1, column=c).fill = navy_fill
ws_dash.row_dimensions[1].height = 36

ws_dash.merge_cells('A2:R2')
ws_dash['A2'].value = "Executive Dashboard -- Key Metrics & Visual Analysis"
ws_dash['A2'].font = h2_font
ws_dash['A2'].fill = lt_blue_fill
for c in range(1, 19):
    ws_dash.cell(row=2, column=c).fill = lt_blue_fill
ws_dash.row_dimensions[2].height = 24

# KPI cards
kpi_defs = [
    ("Overall Experience Index", f"=ROUND('Scoring Engine'!B11,1)", GREEN),
    ("Retention Risk",           f"=ROUND('Scoring Engine'!B12,1)", BLUE),
    ("Engagement Score",         f"=ROUND(('Scoring Engine'!E5+'Scoring Engine'!E6+'Scoring Engine'!E7+'Scoring Engine'!E8+'Scoring Engine'!E9+'Scoring Engine'!E10)/6,1)", TEAL),
    ("Largest Gap",              "INDEX('Scoring Engine'!A5:A10,MATCH(MAX('Scoring Engine'!D5:D10),'Scoring Engine'!D5:D10,0))", ORANGE),
    ("Strongest Dimension",      "INDEX('Scoring Engine'!A5:A10,MATCH(MAX('Scoring Engine'!E5:E10),'Scoring Engine'!E5:E10,0))", GREEN),
    ("Respondents",              f"=COUNTA('Raw Data'!A4:A{4+NUM_RESPONDENTS-1})", NAVY),
]

for idx, (label, formula, fill_color) in enumerate(kpi_defs):
    col_start = 1 + idx * 3
    col_end = col_start + 2

    ws_dash.merge_cells(start_row=4, start_column=col_start,
                        end_row=6, end_column=col_end)
    cell = ws_dash.cell(row=4, column=col_start)

    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    cell.value = formula
    cell.font = kpi_value_font
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='medium', color=WHITE),
                          right=Side(style='medium', color=WHITE),
                          top=Side(style='medium', color=WHITE),
                          bottom=Side(style='medium', color=WHITE))

    ws_dash.merge_cells(start_row=3, start_column=col_start,
                        end_row=3, end_column=col_end)
    label_cell = ws_dash.cell(row=3, column=col_start)
    label_cell.value = label
    label_cell.font = Font(name='Calibri', bold=True, size=9, color=fill_color)
    label_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type='solid')
    label_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws_dash.row_dimensions[3].height = 18
    ws_dash.row_dimensions[4].height = 28
    ws_dash.row_dimensions[5].height = 28
    ws_dash.row_dimensions[6].height = 28

for i in range(1, 19):
    ws_dash.column_dimensions[get_column_letter(i)].width = 16

for r in [3, 4, 5, 6]:
    for c in range(1, 19):
        cell = ws_dash.cell(row=r, column=c)
        if not cell.fill or str(cell.fill.start_color.rgb) == '00000000':
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type='solid')

# Radar Chart
ws_dash.cell(row=8, column=1, value="Experience Profile -- Radar Chart").font = h2_font
ws_dash.merge_cells('A8:F8')

radar_headers = ["Dimension", "Promise Avg", "Experience Avg"]
for c, h in enumerate(radar_headers, 1):
    ws_dash.cell(row=9, column=c, value=h)
apply_header_style(ws_dash, 9, 3)

for d_idx, dim in enumerate(DIMENSIONS):
    row = 10 + d_idx
    ws_dash.cell(row=row, column=1, value=dim["name"])
    se_row = 5 + d_idx
    ws_dash.cell(row=row, column=2, value=f"='Scoring Engine'!B{se_row}")
    ws_dash.cell(row=row, column=2).number_format = '0.00'
    ws_dash.cell(row=row, column=3, value=f"='Scoring Engine'!C{se_row}")
    ws_dash.cell(row=row, column=3).number_format = '0.00'
    for c in range(1, 4):
        ws_dash.cell(row=row, column=c).font = data_font
        ws_dash.cell(row=row, column=c).alignment = center_align
        ws_dash.cell(row=row, column=c).border = thin_border

radar = RadarChart()
radar.type = "filled"
radar.style = 26
radar.title = "Experience Profile: Promise vs Actual"
radar.width = 18
radar.height = 14

cats = Reference(ws_dash, min_col=1, min_row=10, max_row=15)

p_data = Reference(ws_dash, min_col=2, min_row=9, max_row=15)
radar.add_data(p_data, titles_from_data=True)
radar.set_categories(cats)

e_data = Reference(ws_dash, min_col=3, min_row=9, max_row=15)
radar.add_data(e_data, titles_from_data=True)

s0 = radar.series[0]
s0.graphicalProperties.solidFill = "2E75B6"
s0.graphicalProperties.line.solidFill = "1F4E79"

s1 = radar.series[1]
s1.graphicalProperties.solidFill = "ED7D31"
s1.graphicalProperties.line.solidFill = "C00000"

radar.legend.position = 'b'
ws_dash.add_chart(radar, "E8")

# Heat Map
heat_start_row = 9
heat_start_col = 7

ws_dash.cell(row=8, column=heat_start_col,
             value="Gap Analysis Heat Map -- All Items").font = h2_font
ws_dash.merge_cells(start_row=8, start_column=heat_start_col,
                    end_row=8, end_column=heat_start_col + 5)

heat_headers = ["Dimension"]
for item_n in range(1, 6):
    heat_headers.append(f"Item {item_n}")
for c, h in enumerate(heat_headers, heat_start_col):
    ws_dash.cell(row=heat_start_row, column=c, value=h)
apply_header_style(ws_dash, heat_start_row, heat_start_col + 5)

for d_idx, dim in enumerate(DIMENSIONS):
    row = heat_start_row + 1 + d_idx
    ws_dash.cell(row=row, column=heat_start_col, value=dim["name"])
    ws_dash.cell(row=row, column=heat_start_col).font = bold_data_font

    for item_idx in range(5):
        global_i = d_idx * 5 + item_idx
        p_col = raw_data_col_letter(global_i, True)
        e_col = raw_data_col_letter(global_i, False)
        col = heat_start_col + 1 + item_idx
        ws_dash.cell(row=row, column=col,
                     value=f"=AVERAGE('Raw Data'!{p_col}4:{p_col}{4+NUM_RESPONDENTS-1})"
                           f"-AVERAGE('Raw Data'!{e_col}4:{e_col}{4+NUM_RESPONDENTS-1})")
        ws_dash.cell(row=row, column=col).number_format = '0.00'
        ws_dash.cell(row=row, column=col).font = data_font
        ws_dash.cell(row=row, column=col).alignment = center_align
        ws_dash.cell(row=row, column=col).border = thin_border

hs_col = get_column_letter(heat_start_col + 1)
he_col = get_column_letter(heat_start_col + 5)
ws_dash.conditional_formatting.add(
    f'{hs_col}{heat_start_row+1}:{he_col}{heat_start_row+6}',
    ColorScaleRule(
        start_type='num', start_value=0,   start_color='00B050',
        mid_type='num',   mid_value=1.5,    mid_color='FFC000',
        end_type='num',   end_value=3.5,    end_color='C00000'
    )
)

legend_row = heat_start_row + 8
ws_dash.merge_cells(start_row=legend_row, start_column=heat_start_col,
                    end_row=legend_row, end_column=heat_start_col+5)
ws_dash.cell(row=legend_row, column=heat_start_col,
    value="<= 1.0 Healthy  |  1.0-2.0 Watch  |  2.0-3.0 Concern  |  >= 3.0 High Risk").font = small_font

# Risk Matrix
risk_matrix_row = 18
ws_dash.cell(row=risk_matrix_row, column=1,
             value="Risk Matrix -- Gap vs Importance").font = h2_font
ws_dash.merge_cells(f'A{risk_matrix_row}:F{risk_matrix_row}')

rm_start = risk_matrix_row + 1
rm_headers = ["Dimension", "Gap Size", "Weight", "Risk Score"]
for c, h in enumerate(rm_headers, 1):
    ws_dash.cell(row=rm_start, column=c, value=h)
apply_header_style(ws_dash, rm_start, 4)

for d_idx, dim in enumerate(DIMENSIONS):
    row = rm_start + 1 + d_idx
    se_row = 5 + d_idx
    ws_dash.cell(row=row, column=1, value=dim["name"])
    ws_dash.cell(row=row, column=2, value=f"='Scoring Engine'!D{se_row}")
    ws_dash.cell(row=row, column=2).number_format = '0.00'
    ws_dash.cell(row=row, column=3, value=dim["weight"])
    ws_dash.cell(row=row, column=3).number_format = '0%'
    ws_dash.cell(row=row, column=4,
                 value=f"=B{row}*C{row}*100")
    ws_dash.cell(row=row, column=4).number_format = '0.0'
    for c in range(1, 5):
        ws_dash.cell(row=row, column=c).font = data_font
        ws_dash.cell(row=row, column=c).alignment = center_align
        ws_dash.cell(row=row, column=c).border = thin_border

rm_chart = ScatterChart()
rm_chart.title = "Risk Matrix -- Priority Quadrants"
rm_chart.x_axis.title = "Gap Size"
rm_chart.y_axis.title = "Weight x 100"
rm_chart.width = 16
rm_chart.height = 12
rm_chart.x_axis.scaling.min = 0
rm_chart.x_axis.scaling.max = 5
rm_chart.y_axis.scaling.min = 0
rm_chart.y_axis.scaling.max = 25

for d_idx, dim in enumerate(DIMENSIONS):
    row = rm_start + 1 + d_idx
    ws_dash.cell(row=row, column=5, value=dim["weight"] * 100)
    ws_dash.cell(row=row, column=5).number_format = '0.0'

y_vals = Reference(ws_dash, min_col=5, min_row=rm_start+1, max_row=rm_start+6)
x_vals = Reference(ws_dash, min_col=2, min_row=rm_start+1, max_row=rm_start+6)

ws_dash.cell(row=rm_start+1, column=6, value=f"=A{rm_start+1}")
ws_dash.cell(row=rm_start+2, column=6, value=f"=A{rm_start+2}")
ws_dash.cell(row=rm_start+3, column=6, value=f"=A{rm_start+3}")
ws_dash.cell(row=rm_start+4, column=6, value=f"=A{rm_start+4}")
ws_dash.cell(row=rm_start+5, column=6, value=f"=A{rm_start+5}")
ws_dash.cell(row=rm_start+6, column=6, value=f"=A{rm_start+6}")

labels = Reference(ws_dash, min_col=6, min_row=rm_start+1, max_row=rm_start+6)

s = Series(y_vals, x_vals, title="Dimensions")
s.graphicalProperties.line.noFill = True
s.marker.symbol = 'circle'
s.marker.size = 8
s.marker.graphicalProperties.solidFill = "1F4E79"
s.dLbls = DataLabelList()
s.dLbls.showVal = False
s.dLbls.showCatName = True
rm_chart.series.append(s)

# Quadrant divider lines
ws_dash.cell(row=rm_start+1, column=7, value=2.0)
ws_dash.cell(row=rm_start+2, column=7, value=2.0)
ws_dash.cell(row=rm_start+1, column=8, value=0)
ws_dash.cell(row=rm_start+2, column=8, value=25)

ws_dash.cell(row=rm_start+1, column=9, value=0)
ws_dash.cell(row=rm_start+2, column=9, value=5)
ws_dash.cell(row=rm_start+1, column=10, value=15)
ws_dash.cell(row=rm_start+2, column=10, value=15)

ws_dash.add_chart(rm_chart, f"A{rm_start + 9}")

# Department Comparison
dept_comp_col = 11
ws_dash.cell(row=risk_matrix_row, column=dept_comp_col,
             value="Department Comparison (Filterable)").font = h2_font
ws_dash.merge_cells(start_row=risk_matrix_row,
                    start_column=dept_comp_col,
                    end_row=risk_matrix_row,
                    end_column=dept_comp_col + 6)

filt_row = risk_matrix_row + 1
ws_dash.cell(row=filt_row, column=dept_comp_col,
             value="Filter: Department").font = h3_font
ws_dash.cell(row=filt_row+1, column=dept_comp_col,
             value="All").font = bold_data_font

dv_dept = DataValidation(type="list",
                         formula1=f'"All,{",".join(DEPARTMENTS)}"',
                         allow_blank=True)
dv_dept.prompt = "Select a department to filter"
dv_dept.promptTitle = "Department Filter"
dv_dept.add(ws_dash.cell(row=filt_row+1, column=dept_comp_col))
ws_dash.add_data_validation(dv_dept)

comp_start = filt_row + 3
comp_headers = ["Dimension", "Avg Promise", "Avg Experience", "Gap", "Health Score"]
for c, h in enumerate(comp_headers, dept_comp_col):
    ws_dash.cell(row=comp_start, column=c, value=h)
apply_header_style(ws_dash, comp_start, dept_comp_col + 4)

filter_cell = f"${get_column_letter(dept_comp_col)}${filt_row+1}"

for d_idx, dim in enumerate(DIMENSIONS):
    row = comp_start + 1 + d_idx
    ws_dash.cell(row=row, column=dept_comp_col, value=dim["name"])
    ws_dash.cell(row=row, column=dept_comp_col).font = bold_data_font

    p_refs_if = []
    e_refs_if = []
    for item_idx in range(len(dim["items"])):
        global_i = d_idx * 5 + item_idx
        p_col = raw_data_col_letter(global_i, True)
        e_col = raw_data_col_letter(global_i, False)

        rd_dept_range = f"'Raw Data'!$B$4:$B${4+NUM_RESPONDENTS-1}"
        rd_p_range = f"'Raw Data'!{p_col}$4:{p_col}${4+NUM_RESPONDENTS-1}"
        rd_e_range = f"'Raw Data'!{e_col}$4:{e_col}${4+NUM_RESPONDENTS-1}"

        p_refs_if.append(
            f"IF({filter_cell}=\"All\","
            f"AVERAGE({rd_p_range}),"
            f"AVERAGEIF({rd_dept_range},{filter_cell},{rd_p_range}))")
        e_refs_if.append(
            f"IF({filter_cell}=\"All\","
            f"AVERAGE({rd_e_range}),"
            f"AVERAGEIF({rd_dept_range},{filter_cell},{rd_e_range}))")

    ws_dash.cell(row=row, column=dept_comp_col+1,
                 value=f"=({'+'.join(p_refs_if)})/{len(dim['items'])}")
    ws_dash.cell(row=row, column=dept_comp_col+1).number_format = '0.00'

    ws_dash.cell(row=row, column=dept_comp_col+2,
                 value=f"=({'+'.join(e_refs_if)})/{len(dim['items'])}")
    ws_dash.cell(row=row, column=dept_comp_col+2).number_format = '0.00'

    ws_dash.cell(row=row, column=dept_comp_col+3,
                 value=f"={get_column_letter(dept_comp_col+1)}{row}"
                       f"-{get_column_letter(dept_comp_col+2)}{row}")
    ws_dash.cell(row=row, column=dept_comp_col+3).number_format = '0.00'

    ws_dash.cell(row=row, column=dept_comp_col+4,
                 value=f"=MAX(0,MIN(100,100-{get_column_letter(dept_comp_col+3)}{row}*20))")
    ws_dash.cell(row=row, column=dept_comp_col+4).number_format = '0.0'

    for c in range(dept_comp_col, dept_comp_col + 5):
        ws_dash.cell(row=row, column=c).font = data_font
        ws_dash.cell(row=row, column=c).alignment = center_align
        ws_dash.cell(row=row, column=c).border = thin_border

loc_filt_row = filt_row + 9
ws_dash.cell(row=loc_filt_row, column=dept_comp_col,
             value="Filter: Location").font = h3_font
ws_dash.cell(row=loc_filt_row+1, column=dept_comp_col,
             value="All").font = bold_data_font

dv_loc = DataValidation(type="list",
                         formula1=f'"All,{",".join(LOCATIONS)}"',
                         allow_blank=True)
dv_loc.prompt = "Select a location to filter"
dv_loc.promptTitle = "Location Filter"
dv_loc.add(ws_dash.cell(row=loc_filt_row+1, column=dept_comp_col))
ws_dash.add_data_validation(dv_loc)

dp_col = get_column_letter(dept_comp_col + 3)
ws_dash.conditional_formatting.add(
    f'{dp_col}{comp_start+1}:{dp_col}{comp_start+6}',
    ColorScaleRule(
        start_type='num', start_value=0,   start_color='00B050',
        mid_type='num',   mid_value=1.5,    mid_color='FFC000',
        end_type='num',   end_value=3.5,    end_color='C00000'
    )
)

for col_letter in ['A','B','C','D','E','F']:
    ws_dash.column_dimensions[col_letter].width = 16
for col_letter in ['G','H','I','J']:
    ws_dash.column_dimensions[col_letter].width = 10
for col_letter in ['K','L','M','N','O','P','Q','R']:
    ws_dash.column_dimensions[col_letter].width = 14

ws_dash.sheet_view.showGridLines = False
ws_dash.freeze_panes = 'A7'

print("[OK] Sheet 4: Dashboard -- complete")


# ===== SHEET 5: RELIABILITY ANALYSIS =======================
ws_rel = wb.create_sheet("Reliability Analysis")
ws_rel.sheet_properties.tabColor = "7030A0"

ws_rel.merge_cells('A1:H1')
ws_rel['A1'].value = "Cronbach's Alpha -- Internal Consistency Reliability"
ws_rel['A1'].font = h1_font
ws_rel.row_dimensions[1].height = 28

ws_rel.merge_cells('A2:H2')
ws_rel['A2'].value = "Target: alpha >= 0.70. All calculations use VAR.S and native Excel formulas."
ws_rel['A2'].font = Font(name='Calibri', italic=True, color=DARK_GRAY, size=9)

rel_header_row = 4
rel_headers = ["Dimension", "# Items (k)", "Sum Item Variances",
               "Total Score Variance", "Cronbach's alpha", "Interpretation"]
for c, h in enumerate(rel_headers, 1):
    ws_rel.cell(row=rel_header_row, column=c, value=h)
apply_header_style(ws_rel, rel_header_row, 6)

for d_idx, dim in enumerate(DIMENSIONS):
    row = rel_header_row + 1 + d_idx
    k = len(dim["items"])

    ws_rel.cell(row=row, column=1, value=dim["name"])
    ws_rel.cell(row=row, column=2, value=k)

    var_refs = []
    for item_idx in range(k):
        global_i = d_idx * 5 + item_idx
        e_col = raw_data_col_letter(global_i, False)
        item_range = f"'Raw Data'!{e_col}4:{e_col}{4+NUM_RESPONDENTS-1}"
        var_refs.append(f"VAR.S({item_range})")

    ws_rel.cell(row=row, column=3,
                value=f"=SUM({','.join(var_refs)})")
    ws_rel.cell(row=row, column=3).number_format = '0.000'

    helper_col = 7
    for resp_offset in range(NUM_RESPONDENTS):
        rd_row = 4 + resp_offset
        sum_cells = []
        for item_idx in range(k):
            global_i = d_idx * 5 + item_idx
            e_col = raw_data_col_letter(global_i, False)
            sum_cells.append(f"'Raw Data'!{e_col}{rd_row}")
        hr = row + 100 + resp_offset
        ws_rel.cell(row=hr, column=helper_col,
                    value=f"=SUM({'+'.join(sum_cells)})")

    helper_col_letter = get_column_letter(helper_col)
    total_range = f"{helper_col_letter}{row+100}:{helper_col_letter}{row+100+NUM_RESPONDENTS-1}"
    ws_rel.cell(row=row, column=4,
                value=f"=VAR.S({total_range})")
    ws_rel.cell(row=row, column=4).number_format = '0.000'

    ws_rel.cell(row=row, column=5,
                value=f"=(B{row}/(B{row}-1))*(1-C{row}/D{row})")
    ws_rel.cell(row=row, column=5).number_format = '0.000'

    ws_rel.cell(row=row, column=6,
                value=f'=IF(E{row}>=0.9,"Excellent",'
                       f'IF(E{row}>=0.8,"Good",'
                       f'IF(E{row}>=0.7,"Acceptable",'
                       f'IF(E{row}>=0.6,"Questionable",'
                       f'IF(E{row}>=0.5,"Poor","Unacceptable")))))')

    for c in range(1, 7):
        ws_rel.cell(row=row, column=c).font = data_font
        ws_rel.cell(row=row, column=c).alignment = center_align
        ws_rel.cell(row=row, column=c).border = thin_border
        if row % 2 == 0:
            ws_rel.cell(row=row, column=c).fill = lt_gray_fill

ws_rel.conditional_formatting.add(
    f'E{rel_header_row+1}:E{rel_header_row+6}',
    ColorScaleRule(
        start_type='num', start_value=0.5,   start_color='C00000',
        mid_type='num',   mid_value=0.7,      mid_color='FFC000',
        end_type='num',   end_value=0.95,     end_color='00B050'
    )
)

sum_row = rel_header_row + 8
ws_rel.merge_cells(f'A{sum_row}:F{sum_row}')
ws_rel.cell(row=sum_row, column=1,
    value="Cronbach's Alpha measures internal consistency: how closely related the items are as a group. "
          "alpha >= 0.70 is the research standard for acceptable reliability.").font = small_font

set_col_widths(ws_rel, {'A': 24, 'B': 12, 'C': 18, 'D': 20, 'E': 16, 'F': 18})
ws_rel.freeze_panes = f'A{rel_header_row+1}'

print("[OK] Sheet 5: Reliability Analysis -- complete")


# ===== SHEET 6: DRIVER ANALYSIS ============================
ws_driver = wb.create_sheet("Driver Analysis")
ws_driver.sheet_properties.tabColor = TEAL

ws_driver.merge_cells('A1:H1')
ws_driver['A1'].value = "Driver Analysis -- Correlation with Overall Experience Index"
ws_driver['A1'].font = h1_font
ws_driver.row_dimensions[1].height = 28

ws_driver.merge_cells('A2:H2')
ws_driver['A2'].value = "CORREL() function measures the strength of each dimension's relationship with the Overall Index."
ws_driver['A2'].font = Font(name='Calibri', italic=True, color=DARK_GRAY, size=9)

corr_header_row = 4
corr_headers = ["Dimension", "Correlation with Overall Index",
                "Strength", "Rank (1=Strongest)"]
for c, h in enumerate(corr_headers, 1):
    ws_driver.cell(row=corr_header_row, column=c, value=h)
apply_header_style(ws_driver, corr_header_row, 4)

# Per-respondent data starts at row 17 = b_header_row + 1 = 16 + 1
pr_data_start = 17
pr_last_row = pr_data_start + NUM_RESPONDENTS - 1

# Overall_E is at column 25 (after 5 demo + 6 dims * 3 cols = 23 + Overall_P(24) + Overall_E(25))
overall_e_col = get_column_letter(25)

# Dimension Experience columns
dim_e_cols = []
for d_idx in range(6):
    dim_e_cols.append(get_column_letter(6 + d_idx * 3 + 1))

for d_idx, dim in enumerate(DIMENSIONS):
    row = corr_header_row + 1 + d_idx
    ws_driver.cell(row=row, column=1, value=dim["name"])

    dim_range = f"'Scoring Engine'!{dim_e_cols[d_idx]}{pr_data_start}:{dim_e_cols[d_idx]}{pr_last_row}"
    overall_range = f"'Scoring Engine'!{overall_e_col}{pr_data_start}:{overall_e_col}{pr_last_row}"

    ws_driver.cell(row=row, column=2,
                   value=f"=CORREL({dim_range},{overall_range})")
    ws_driver.cell(row=row, column=2).number_format = '0.000'

    ws_driver.cell(row=row, column=3,
                   value=f'=IF(ABS(B{row})>=0.7,"Strong",'
                          f'IF(ABS(B{row})>=0.5,"Moderate",'
                          f'IF(ABS(B{row})>=0.3,"Weak","Negligible")))')

    ws_driver.cell(row=row, column=4,
                   value=f'=IF(B{row}<>"",RANK(B{row},$B${corr_header_row+1}:$B${corr_header_row+6}),"")')

    for c in range(1, 5):
        ws_driver.cell(row=row, column=c).font = data_font
        ws_driver.cell(row=row, column=c).alignment = center_align
        ws_driver.cell(row=row, column=c).border = thin_border
        if row % 2 == 0:
            ws_driver.cell(row=row, column=c).fill = lt_gray_fill

set_col_widths(ws_driver, {'A': 24, 'B': 32, 'C': 16, 'D': 22})

exp_row = corr_header_row + 9
ws_driver.merge_cells(f'A{exp_row}:D{exp_row}')
ws_driver.cell(row=exp_row, column=1,
    value="Driver Analysis interprets CORREL() results: values near +1 = strong positive relationship; "
          "near 0 = no relationship. The higher the correlation, the more that dimension drives "
          "the overall experience score.").font = small_font
exp_row += 1
ws_driver.merge_cells(f'A{exp_row}:D{exp_row}')
ws_driver.cell(row=exp_row, column=1,
    value="Action: Focus improvement efforts on dimensions with the strongest correlation "
          "AND the largest gap -- these are 'high-impact' drivers.").font = small_font

bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.title = "Driver Strength -- Correlation with Overall Experience"
bar_chart.y_axis.title = "Correlation (r)"
bar_chart.style = 10
bar_chart.width = 20
bar_chart.height = 12

data = Reference(ws_driver, min_col=2, min_row=corr_header_row,
                 max_row=corr_header_row+6, max_col=2)
cats = Reference(ws_driver, min_col=1, min_row=corr_header_row+1,
                 max_row=corr_header_row+6)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
bar_chart.shape = 4

s0 = bar_chart.series[0]
s0.graphicalProperties.solidFill = "2E75B6"

ws_driver.add_chart(bar_chart, f"A{exp_row + 2}")
ws_driver.freeze_panes = f'A{corr_header_row+1}'

print("[OK] Sheet 6: Driver Analysis -- complete")


# ===== SHEET 7: TURNOVER PREDICTION ========================
ws_turn = wb.create_sheet("Turnover Prediction")
ws_turn.sheet_properties.tabColor = RED

ws_turn.merge_cells('A1:H1')
ws_turn['A1'].value = "Turnover Prediction Model -- Retention Risk Assessment"
ws_turn['A1'].font = h1_font
ws_turn.row_dimensions[1].height = 28

ws_turn.merge_cells('A2:H2')
ws_turn['A2'].value = "Model: Turnover Probability = (Retention Risk x 0.5) + (Manager Gap x 0.3) + (Culture Gap x 0.2)"
ws_turn['A2'].font = Font(name='Calibri', italic=True, color=DARK_GRAY, size=9)

agg_row = 4
ws_turn.cell(row=agg_row, column=1,
             value="Aggregate Turnover Prediction").font = h2_font
ws_turn.merge_cells(f'A{agg_row}:F{agg_row}')

metrics = [
    ("Retention Risk Score", "='Scoring Engine'!B12", "0.0"),
    ("Manager Gap (weight 30%)", "='Scoring Engine'!D6", "0.00"),
    ("Culture Gap (weight 20%)", "='Scoring Engine'!D8", "0.00"),
    ("Turnover Probability", "=(B5*0.5)+(B6*0.3)+(B7*0.2)", "0.0%"),
    ("Risk Classification", '=IF(B8<=0.2,"Very Low",IF(B8<=0.4,"Low",IF(B8<=0.6,"Moderate",IF(B8<=0.8,"High","Critical"))))', ""),
]

for idx, (label, formula, fmt) in enumerate(metrics):
    row = agg_row + 1 + idx
    ws_turn.cell(row=row, column=1, value=label).font = bold_data_font
    ws_turn.cell(row=row, column=2, value=formula)
    color = RED if ('Risk' in label or 'Turnover' in label) else DARK_GRAY
    ws_turn.cell(row=row, column=2).font = Font(name='Calibri', bold=True, size=12, color=color)
    ws_turn.cell(row=row, column=2).number_format = fmt
    ws_turn.cell(row=row, column=2).alignment = center_align
    fill_c = lt_red_fill if idx == 3 else lt_gray_fill
    for c in [1, 2]:
        ws_turn.cell(row=row, column=c).fill = fill_c
        ws_turn.cell(row=row, column=c).border = thin_border

turn_header_row = agg_row + 8
ws_turn.merge_cells(f'A{turn_header_row}:H{turn_header_row}')
ws_turn.cell(row=turn_header_row, column=1,
             value="Per-Respondent Turnover Risk").font = h2_font

t_headers = ["RespID", "Department", "Manager Gap", "Culture Gap",
             "Retention Risk", "Turnover Probability", "Risk Level"]
for c, h in enumerate(t_headers, 1):
    ws_turn.cell(row=turn_header_row+1, column=c, value=h)
apply_header_style(ws_turn, turn_header_row+1, len(t_headers))

# Columns in Scoring Engine per-respondent section
# 5 demo cols + 6 dims*3 = 23, then Overall_P=24, Overall_E=25, Overall_Gap=26, Risk=27, Turnover=28
resp_id_col = "A"
dept_col    = "B"
me_gap_col  = get_column_letter(11)
wc_gap_col  = get_column_letter(17)
risk_col    = get_column_letter(27)
turn_col    = get_column_letter(28)

for resp_offset in range(NUM_RESPONDENTS):
    row = turn_header_row + 2 + resp_offset
    se_row = pr_data_start + resp_offset

    ws_turn.cell(row=row, column=1,
                 value=f"='Scoring Engine'!{resp_id_col}{se_row}")
    ws_turn.cell(row=row, column=2,
                 value=f"='Scoring Engine'!{dept_col}{se_row}")
    ws_turn.cell(row=row, column=3,
                 value=f"='Scoring Engine'!{me_gap_col}{se_row}")
    ws_turn.cell(row=row, column=3).number_format = '0.00'
    ws_turn.cell(row=row, column=4,
                 value=f"='Scoring Engine'!{wc_gap_col}{se_row}")
    ws_turn.cell(row=row, column=4).number_format = '0.00'
    ws_turn.cell(row=row, column=5,
                 value=f"='Scoring Engine'!{risk_col}{se_row}")
    ws_turn.cell(row=row, column=5).number_format = '0.00'
    ws_turn.cell(row=row, column=6,
                 value=f"='Scoring Engine'!{turn_col}{se_row}")
    ws_turn.cell(row=row, column=6).number_format = '0.0%'
    ws_turn.cell(row=row, column=7,
                 value=f'=IF(F{row}<=0.2,"Very Low",'
                        f'IF(F{row}<=0.4,"Low",'
                        f'IF(F{row}<=0.6,"Moderate",'
                        f'IF(F{row}<=0.8,"High","Critical"))))')

    for c in range(1, 8):
        ws_turn.cell(row=row, column=c).font = data_font
        ws_turn.cell(row=row, column=c).alignment = center_align
        ws_turn.cell(row=row, column=c).border = thin_border
        if row % 2 == 0:
            ws_turn.cell(row=row, column=c).fill = lt_gray_fill

# Conditional formatting for turnover probability
turn_data_start = turn_header_row + 2
turn_data_end = turn_header_row + 1 + NUM_RESPONDENTS

ws_turn.conditional_formatting.add(
    f'F{turn_data_start}:F{turn_data_end}',
    ColorScaleRule(
        start_type='num', start_value=0,     start_color='00B050',
        mid_type='num',   mid_value=0.4,      mid_color='FFC000',
        end_type='num',   end_value=0.8,      end_color='C00000'
    )
)

risk_level_rng = f'G{turn_data_start}:G{turn_data_end}'
ws_turn.conditional_formatting.add(
    risk_level_rng,
    CellIsRule(operator='equal', formula=['"Very Low"'],
               fill=lt_green_fill, font=Font(color=GREEN, bold=True))
)
ws_turn.conditional_formatting.add(
    risk_level_rng,
    CellIsRule(operator='equal', formula=['"Low"'],
               fill=lt_green_fill, font=Font(color=GREEN, bold=True))
)
ws_turn.conditional_formatting.add(
    risk_level_rng,
    CellIsRule(operator='equal', formula=['"Moderate"'],
               fill=lt_amber_fill, font=Font(color=AMBER, bold=True))
)
ws_turn.conditional_formatting.add(
    risk_level_rng,
    CellIsRule(operator='equal', formula=['"High"'],
               fill=lt_orange_fill, font=Font(color=ORANGE, bold=True))
)
ws_turn.conditional_formatting.add(
    risk_level_rng,
    CellIsRule(operator='equal', formula=['"Critical"'],
               fill=lt_red_fill, font=Font(color=RED, bold=True))
)

stat_row = turn_data_end + 2
ws_turn.merge_cells(f'A{stat_row}:G{stat_row}')
ws_turn.cell(row=stat_row, column=1,
             value="Summary Statistics").font = h2_font

stats_list = [
    ("Average Turnover Probability", f"=AVERAGE(F{turn_data_start}:F{turn_data_end})", '0.0%', RED),
    ("High Risk Employees (>=60%)",  f"=COUNTIF(F{turn_data_start}:F{turn_data_end},\">=0.6\")", '0', RED),
    ("Moderate Risk (40-60%)",       f"=COUNTIFS(F{turn_data_start}:F{turn_data_end},\">=0.4\",F{turn_data_start}:F{turn_data_end},\"<0.6\")", '0', ORANGE),
    ("Low Risk (<40%)",              f"=COUNTIF(F{turn_data_start}:F{turn_data_end},\"<0.4\")", '0', GREEN),
]

for idx, (label, formula, fmt, color) in enumerate(stats_list):
    row = stat_row + 1 + idx
    ws_turn.cell(row=row, column=1, value=label).font = bold_data_font
    ws_turn.cell(row=row, column=2, value=formula)
    ws_turn.cell(row=row, column=2).font = Font(name='Calibri', bold=True, size=12, color=color)
    ws_turn.cell(row=row, column=2).number_format = fmt
    ws_turn.cell(row=row, column=2).alignment = center_align
    for c in [1, 2]:
        ws_turn.cell(row=row, column=c).border = thin_border
        ws_turn.cell(row=row, column=c).fill = lt_gray_fill

set_col_widths(ws_turn, {'A': 12, 'B': 20, 'C': 15, 'D': 15, 'E': 16, 'F': 20, 'G': 14})
ws_turn.freeze_panes = f'A{turn_data_start}'

print("[OK] Sheet 7: Turnover Prediction -- complete")


# ============================================================
# FINAL FORMATTING & SAVE
# ============================================================
for ws in [ws_survey, ws_data, ws_score, ws_dash, ws_rel, ws_driver, ws_turn]:
    ws.sheet_view.showGridLines = ws.title in ("Raw Data", "Survey")
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A3

# Save to temp (guaranteed writable), then copy to project folder
wb.save(TEMP_FILE)
if os.path.exists(OUTPUT_FILE):
    try:
        os.remove(OUTPUT_FILE)
    except PermissionError:
        pass  # File open in Excel, copy will overwrite below
shutil.copy2(TEMP_FILE, OUTPUT_FILE)
print(f"[OK] Workbook saved to project folder: {OUTPUT_FILE}")
print(f"\nSheets created:")
print(f"  1. Survey           -- 30 items, 6 dimensions, dual-scale")
print(f"  2. Raw Data         -- {NUM_RESPONDENTS} sample respondents")
print(f"  3. Scoring Engine   -- Formulas: gaps, health scores, overall index, risk")
print(f"  4. Dashboard        -- KPIs, Radar Chart, Heat Map, Risk Matrix, Filters")
print(f"  5. Reliability      -- Cronbach's Alpha per dimension")
print(f"  6. Driver Analysis  -- CORREL() with bar chart")
print(f"  7. Turnover Predict -- Retention risk + turnover probability model")
print(f"\nNative Excel formulas used throughout. No VBA required.")
print(f"Add more respondents by extending rows in 'Raw Data' sheet.")
